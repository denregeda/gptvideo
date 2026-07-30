"""Отчёты, конструктор отчётов и экспорт в Excel/PDF."""
import io
import json
import logging
from datetime import datetime, timezone, timedelta
from typing import Optional

from fastapi import APIRouter, Body, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy import text
from sqlalchemy.orm import Session

from deps import get_db, get_current_admin

log = logging.getLogger(__name__)

DOW_RU = ["Пн", "Вт", "Ср", "Чт", "Пт", "Сб", "Вс"]

router = APIRouter()


@router.get("/reports/summary")
def report_summary(admin: dict = Depends(get_current_admin), db: Session = Depends(get_db)):
    total = db.execute(text("SELECT COUNT(*) FROM screens")).scalar()
    online = db.execute(text(
        "SELECT COUNT(*) FROM screens WHERE last_seen > NOW() - INTERVAL '5 minutes'"
    )).scalar()
    media_count = db.execute(text("SELECT COUNT(*) FROM media WHERE status = 'ready'")).scalar()
    return {"total_screens": total, "online": online, "offline": total - online,
            "media_files": media_count,
            "generated_at": datetime.now(timezone.utc).isoformat()}


@router.get("/reports/screen/{screen_id}")
def report_screen(screen_id: int, days: int = Query(7), admin: dict = Depends(get_current_admin),
                  db: Session = Depends(get_db)):
    # ИСПРАВЛЕНО: правильная параметризация INTERVAL
    rows = db.execute(text("""
        SELECT pl.filename, COUNT(*) as play_count,
               SUM(EXTRACT(EPOCH FROM (pl.ended_at - pl.started_at))) as total_seconds
        FROM play_log pl
        WHERE pl.screen_id = :sid
          AND pl.started_at > NOW() - (INTERVAL '1 day' * :days)
        GROUP BY pl.filename ORDER BY play_count DESC
    """), {"sid": screen_id, "days": days}).fetchall()
    return [dict(r._mapping) for r in rows]


def _data_offline_screens(db: Session):
    rows = db.execute(text("""
        SELECT name, city, last_seen FROM screens
        WHERE last_seen IS NULL OR last_seen < NOW() - INTERVAL '5 minutes'
        ORDER BY last_seen NULLS FIRST
    """)).fetchall()
    return [dict(r._mapping) for r in rows]


@router.get("/reports/offline-screens")
def report_offline_screens(admin: dict = Depends(get_current_admin), db: Session = Depends(get_db)):
    screens = _data_offline_screens(db)
    return {"count": len(screens), "screens": screens}


def _data_low_disk(db: Session, threshold_pct: float = 15):
    rows = db.execute(text("""
        SELECT name, disk_free_gb, disk_total_gb
        FROM screens
        WHERE disk_free_gb IS NOT NULL AND disk_total_gb IS NOT NULL AND disk_total_gb > 0
          AND (disk_free_gb / disk_total_gb * 100) < :threshold
        ORDER BY (disk_free_gb / disk_total_gb) ASC
    """), {"threshold": threshold_pct}).fetchall()
    return [{"name": r.name, "disk_free_gb": r.disk_free_gb, "disk_total_gb": r.disk_total_gb,
             "free_pct": round(r.disk_free_gb / r.disk_total_gb * 100, 1)} for r in rows]


@router.get("/reports/low-disk")
def report_low_disk(threshold_pct: float = Query(15), admin: dict = Depends(get_current_admin),
                     db: Session = Depends(get_db)):
    return _data_low_disk(db, threshold_pct)


def _data_downtime(db: Session, hours_full: float = 8):
    rows = db.execute(text("""
        SELECT name, EXTRACT(EPOCH FROM (NOW() - COALESCE(last_seen, created_at))) / 3600.0 AS hours_offline
        FROM screens
        WHERE last_seen IS NULL OR last_seen < NOW() - INTERVAL '5 minutes'
    """)).fetchall()
    result = []
    for r in rows:
        pct = round(min(100, (float(r.hours_offline) / hours_full) * 100)) if hours_full > 0 else 0
        if pct > 0:
            result.append({"name": r.name, "downtime_pct": pct})
    result.sort(key=lambda x: -x["downtime_pct"])
    return result


@router.get("/reports/downtime")
def report_downtime(hours_full: float = Query(8), admin: dict = Depends(get_current_admin),
                     db: Session = Depends(get_db)):
    return _data_downtime(db, hours_full)


def _data_validity(db: Session, advertiser=None, media_id=None):
    where = ["m.status = 'ready'"]
    params = {}
    if advertiser:
        where.append("a.name = :adv"); params["adv"] = advertiser
    if media_id:
        where.append("m.id = :mid"); params["mid"] = int(media_id)
    rows = db.execute(text(f"""
        SELECT m.title, a.name AS advertiser, m.valid_from, m.valid_until,
               CASE
                 WHEN m.valid_from IS NULL AND m.valid_until IS NULL THEN 'unlimited'
                 WHEN m.valid_from IS NOT NULL AND m.valid_from > NOW() THEN 'upcoming'
                 WHEN m.valid_until IS NOT NULL AND m.valid_until < NOW() THEN 'expired'
                 ELSE 'active'
               END AS validity_status
        FROM media m
        LEFT JOIN advertisers a ON a.id = m.advertiser_id
        WHERE {' AND '.join(where)}
        ORDER BY m.valid_until NULLS LAST, m.title
    """), params).fetchall()
    return [dict(r._mapping) for r in rows]


@router.get("/reports/validity")
def report_validity(advertiser: Optional[str] = Query(None), media_id: Optional[int] = Query(None),
                    admin: dict = Depends(get_current_admin), db: Session = Depends(get_db)):
    return _data_validity(db, advertiser, media_id)


@router.get("/reports/versions")
def report_versions(admin: dict = Depends(get_current_admin), db: Session = Depends(get_db)):
    target = db.execute(text("SELECT os_version, vlc_version FROM target_versions WHERE id = 1")).fetchone()
    target_os = target.os_version if target else None
    target_vlc = target.vlc_version if target else None

    rows = db.execute(text("SELECT name, os_version, vlc_version FROM screens ORDER BY name")).fetchall()
    screens = []
    outdated = 0
    for r in rows:
        os_status = 'ok' if (target_os and r.os_version and r.os_version == target_os) \
            else ('outdated' if (target_os and r.os_version) else 'not_compared')
        vlc_status = 'ok' if (target_vlc and r.vlc_version and r.vlc_version == target_vlc) \
            else ('outdated' if (target_vlc and r.vlc_version) else 'not_compared')
        if os_status == 'outdated' or vlc_status == 'outdated':
            outdated += 1
        screens.append({"name": r.name, "os_version": r.os_version, "os_status": os_status,
                         "vlc_version": r.vlc_version, "vlc_status": vlc_status})

    return {"target_os": target_os, "target_vlc": target_vlc, "outdated_count": outdated, "screens": screens}


def _data_connection_losses(db: Session, days: int = 30):
    gap_rows = db.execute(text("""
        WITH ordered AS (
            SELECT s.name, ms.screen_id, ms.timestamp,
                   LAG(ms.timestamp) OVER (PARTITION BY ms.screen_id ORDER BY ms.timestamp) AS prev_ts
            FROM minipc_status ms
            JOIN screens s ON s.id = ms.screen_id
            WHERE ms.timestamp > NOW() - (INTERVAL '1 day' * :days)
        )
        SELECT name, prev_ts AS lost_at, timestamp AS restored_at,
               EXTRACT(EPOCH FROM (timestamp - prev_ts)) AS seconds
        FROM ordered
        WHERE prev_ts IS NOT NULL AND timestamp - prev_ts > INTERVAL '5 minutes'
        ORDER BY prev_ts DESC
    """), {"days": days}).fetchall()

    ongoing_rows = db.execute(text("""
        SELECT name, last_seen AS lost_at, EXTRACT(EPOCH FROM (NOW() - last_seen)) AS seconds
        FROM screens WHERE last_seen IS NOT NULL AND last_seen < NOW() - INTERVAL '5 minutes'
    """)).fetchall()

    result = [{"screen": r.name, "lost_at": r.lost_at, "restored_at": r.restored_at,
               "seconds": r.seconds, "ongoing": False} for r in gap_rows]
    result += [{"screen": r.name, "lost_at": r.lost_at, "restored_at": None,
                "seconds": r.seconds, "ongoing": True} for r in ongoing_rows]
    result.sort(key=lambda x: x["lost_at"] or "", reverse=True)
    return result


@router.get("/reports/connection-losses")
def report_connection_losses(days: int = Query(30), admin: dict = Depends(get_current_admin),
                              db: Session = Depends(get_db)):
    return _data_connection_losses(db, days)


@router.post("/session/ping")
def session_ping(current_admin: dict = Depends(get_current_admin), db: Session = Depends(get_db)):
    """
    Отметка активности в панели (вызывается фронтендом сразу после логина и раз в 2 минуты).
    Продлевает уже открытую сессию (last_seen) или начинает новую, если предыдущая
    "протухла" (не было пинга больше 5 минут — считаем это новым заходом).
    """
    username = current_admin["username"]
    now = datetime.now(timezone.utc)

    user_row = db.execute(text("SELECT id FROM users WHERE username = :u"), {"u": username}).fetchone()
    user_id = user_row.id if user_row else None

    open_session = db.execute(text("""
        SELECT id FROM user_sessions
        WHERE username = :u AND logout_at IS NULL AND last_seen > :cutoff
        ORDER BY login_at DESC LIMIT 1
    """), {"u": username, "cutoff": now - timedelta(minutes=5)}).fetchone()

    if open_session:
        db.execute(text("UPDATE user_sessions SET last_seen = :now WHERE id = :id"),
                   {"now": now, "id": open_session.id})
    else:
        db.execute(text("""
            INSERT INTO user_sessions (user_id, username, login_at, last_seen)
            VALUES (:uid, :u, :now, :now)
        """), {"uid": user_id, "u": username, "now": now})
    db.commit()
    return {"status": "ok"}


def _data_sessions(db: Session, days: int = 30, date_from=None, date_to=None):
    where, params = [], {}
    if date_from or date_to:
        if date_from:
            where.append("us.login_at >= CAST(:d_from AS date)"); params["d_from"] = date_from
        if date_to:
            where.append("us.login_at < (CAST(:d_to AS date) + 1)"); params["d_to"] = date_to
    else:
        where.append("us.login_at > NOW() - (INTERVAL '1 day' * :days)"); params["days"] = days
    rows = db.execute(text(f"""
        SELECT us.username, COALESCE(u.role, 'admin') AS role, us.login_at,
               EXTRACT(EPOCH FROM (COALESCE(us.logout_at, us.last_seen) - us.login_at)) AS seconds
        FROM user_sessions us
        LEFT JOIN users u ON u.id = us.user_id
        WHERE {' AND '.join(where)}
        ORDER BY us.login_at DESC
    """), params).fetchall()
    return [dict(r._mapping) for r in rows]


@router.get("/reports/sessions")
def report_sessions(days: int = Query(30), date_from: Optional[str] = Query(None),
                     date_to: Optional[str] = Query(None), admin: dict = Depends(get_current_admin),
                     db: Session = Depends(get_db)):
    return _data_sessions(db, days, date_from, date_to)


def _data_broken_media(db: Session, advertiser=None, media_id=None):
    where = ["m.is_broken = TRUE"]
    params = {}
    if advertiser:
        where.append("a.name = :adv"); params["adv"] = advertiser
    if media_id:
        where.append("m.id = :mid"); params["mid"] = int(media_id)
    rows = db.execute(text(f"""
        SELECT m.title, m.filename, a.name AS advertiser, m.error_count, m.last_error
        FROM media m
        LEFT JOIN advertisers a ON a.id = m.advertiser_id
        WHERE {' AND '.join(where)}
        ORDER BY m.last_error_at DESC NULLS LAST
    """), params).fetchall()
    return [dict(r._mapping) for r in rows]


@router.get("/reports/broken-media")
def report_broken_media(advertiser: Optional[str] = Query(None), media_id: Optional[int] = Query(None),
                        admin: dict = Depends(get_current_admin), db: Session = Depends(get_db)):
    return _data_broken_media(db, advertiser, media_id)


# ─── Конструктор отчётов ─────────────────────────────────────────────────────
# Whitelist метрик/разрезов/фильтров — конфиг с фронтенда только ВЫБИРАЕТ ключи
# из этих словарей, произвольный SQL от клиента невозможен в принципе.

BUILDER_METRICS = [
    {"key": "plays", "label": "Показы (количество)"},
    {"key": "seconds", "label": "Время эфира (секунды)"},
]
BUILDER_DIMENSIONS = [
    {"key": "media", "label": "Ролик"},
    {"key": "advertiser", "label": "Рекламодатель"},
    {"key": "screen", "label": "Экран"},
    {"key": "city", "label": "Город"},
    {"key": "day", "label": "День"},
]
BUILDER_PERIODS = [
    {"key": "1", "label": "Сегодня"},
    {"key": "7", "label": "7 дней"},
    {"key": "30", "label": "30 дней"},
    {"key": "90", "label": "90 дней"},
]
BUILDER_FILTERS = [
    {"key": "none", "label": "Без фильтра"},
    {"key": "advertiser", "label": "Рекламодатель"},
    {"key": "media", "label": "Ролик"},
    {"key": "screen", "label": "Экран"},
    {"key": "city", "label": "Город"},
]
BUILDER_SORTS = [
    {"key": "val_desc", "label": "Значение ↓"},
    {"key": "val_asc", "label": "Значение ↑"},
    {"key": "label_asc", "label": "Название А→Я"},
    {"key": "label_desc", "label": "Название Я→А"},
]
BUILDER_LIMITS = [
    {"key": "0", "label": "Без ограничения"},
    {"key": "10", "label": "Топ 10"},
    {"key": "20", "label": "Топ 20"},
    {"key": "50", "label": "Топ 50"},
]

_BUILDER_DIMENSION_SQL = {
    "media": "COALESCE(m.title, pl.filename)",
    "advertiser": "a.name",
    "screen": "s.name",
    "city": "s.city",
    "day": "to_char(pl.started_at + INTERVAL '3 hours', 'YYYY-MM-DD')",
}
_BUILDER_METRIC_SQL = {
    "plays": "COUNT(*)",
    "seconds": "COALESCE(SUM(EXTRACT(EPOCH FROM (pl.ended_at - pl.started_at))), 0)",
}
_BUILDER_METRIC_LABELS = {m["key"]: m["label"] for m in BUILDER_METRICS}
_BUILDER_DIMENSION_LABELS = {d["key"]: d["label"] for d in BUILDER_DIMENSIONS}
_BUILDER_FILTER_SQL = {
    "advertiser": "a.name = :fval",
    "media": "COALESCE(m.title, pl.filename) = :fval",
    "screen": "s.name = :fval",
    "city": "s.city = :fval",
}


@router.get("/reports/builder/options")
def builder_options(admin: dict = Depends(get_current_admin), db: Session = Depends(get_db)):
    advertisers = [r[0] for r in db.execute(text("SELECT name FROM advertisers ORDER BY name")).fetchall()]
    screens = [r[0] for r in db.execute(text("SELECT name FROM screens ORDER BY name")).fetchall()]
    media = [r[0] for r in db.execute(text("""
        SELECT DISTINCT COALESCE(m.title, pl.filename) AS v
        FROM play_log pl LEFT JOIN media m ON m.id = pl.media_id
        WHERE COALESCE(m.title, pl.filename) IS NOT NULL
        ORDER BY v
    """)).fetchall()]
    return {
        "metrics": BUILDER_METRICS, "dimensions": BUILDER_DIMENSIONS,
        "periods": BUILDER_PERIODS, "filters": BUILDER_FILTERS,
        "sorts": BUILDER_SORTS, "limits": BUILDER_LIMITS,
        "advertisers": advertisers, "screens": screens, "media": media,
    }


def _builder_run_data(body: dict, db: Session):
    metric = body.get("metric")
    dimension = body.get("dimension")
    if metric not in _BUILDER_METRIC_SQL:
        raise HTTPException(status_code=400, detail="Недопустимый показатель")
    if dimension not in _BUILDER_DIMENSION_SQL:
        raise HTTPException(status_code=400, detail="Недопустимый разрез")

    days = int(body.get("days") or 7)
    date_from = body.get("date_from")
    date_to = body.get("date_to")
    filter_key = body.get("filter") or "none"
    filter_value = body.get("filter_value")
    sort_by = "val" if body.get("sort_by") != "label" else "label"
    sort_dir = "ASC" if body.get("sort_dir") == "asc" else "DESC"
    limit = int(body.get("limit") or 0)

    # Диапазон дат имеет приоритет над периодом-пресетом (days)
    if date_from or date_to:
        where, params = [], {}
        if date_from:
            where.append("pl.started_at >= CAST(:d_from AS timestamp) - INTERVAL '3 hours'")
            params["d_from"] = date_from
        if date_to:
            where.append("pl.started_at < CAST(:d_to AS timestamp) + INTERVAL '21 hours'")
            params["d_to"] = date_to
    else:
        where = ["pl.started_at > NOW() - (INTERVAL '1 day' * :days)"]
        params = {"days": days}
    if filter_key in _BUILDER_FILTER_SQL and filter_value:
        where.append(_BUILDER_FILTER_SQL[filter_key])
        params["fval"] = filter_value

    limit_sql = f"LIMIT {limit}" if limit and limit > 0 else ""
    sql = f"""
        SELECT {_BUILDER_DIMENSION_SQL[dimension]} AS label, {_BUILDER_METRIC_SQL[metric]} AS val
        FROM play_log pl
        JOIN screens s ON s.id = pl.screen_id
        LEFT JOIN media m ON m.id = pl.media_id
        LEFT JOIN advertisers a ON a.id = m.advertiser_id
        WHERE {' AND '.join(where)}
        GROUP BY {_BUILDER_DIMENSION_SQL[dimension]}
        ORDER BY {sort_by} {sort_dir}
        {limit_sql}
    """
    rows = db.execute(text(sql), params).fetchall()

    title = f"{_BUILDER_METRIC_LABELS[metric]} по: {_BUILDER_DIMENSION_LABELS[dimension]}"
    columns = [_BUILDER_DIMENSION_LABELS[dimension], _BUILDER_METRIC_LABELS[metric]]
    result_rows = [[r.label, round(float(r.val), 2) if r.val is not None else 0] for r in rows]
    return title, columns, result_rows


@router.post("/reports/builder/run")
def builder_run(body: dict = Body(...), admin: dict = Depends(get_current_admin), db: Session = Depends(get_db)):
    title, columns, rows = _builder_run_data(body, db)
    return {"title": title, "columns": columns, "rows": rows}


@router.post("/reports/builder/save")
def builder_save(body: dict = Body(...), admin: dict = Depends(get_current_admin), db: Session = Depends(get_db)):
    name = (body.get("name") or "").strip()
    config = body.get("config") or {}
    if not name:
        raise HTTPException(status_code=400, detail="Название обязательно")

    user_row = db.execute(text("SELECT id FROM users WHERE username = :u"), {"u": admin["username"]}).fetchone()
    owner_id = user_row.id if user_row else None
    db.execute(text("""
        INSERT INTO custom_reports (owner_id, name, config) VALUES (:oid, :name, CAST(:cfg AS jsonb))
    """), {"oid": owner_id, "name": name, "cfg": json.dumps(config)})
    db.commit()
    return {"status": "ok"}


@router.get("/reports/builder/saved")
def builder_saved(admin: dict = Depends(get_current_admin), db: Session = Depends(get_db)):
    user_row = db.execute(text("SELECT id FROM users WHERE username = :u"), {"u": admin["username"]}).fetchone()
    owner_id = user_row.id if user_row else None
    rows = db.execute(text("""
        SELECT id, name, config FROM custom_reports WHERE owner_id = :oid ORDER BY created_at DESC
    """), {"oid": owner_id}).fetchall()
    return [dict(r._mapping) for r in rows]


@router.delete("/reports/builder/saved/{report_id}")
def builder_delete_saved(report_id: int, admin: dict = Depends(get_current_admin), db: Session = Depends(get_db)):
    user_row = db.execute(text("SELECT id FROM users WHERE username = :u"), {"u": admin["username"]}).fetchone()
    owner_id = user_row.id if user_row else None
    row = db.execute(text("SELECT id FROM custom_reports WHERE id = :id AND owner_id = :oid"),
                      {"id": report_id, "oid": owner_id}).fetchone()
    if not row:
        raise HTTPException(status_code=404, detail="Отчёт не найден")
    db.execute(text("DELETE FROM custom_reports WHERE id = :id"), {"id": report_id})
    db.commit()
    return {"status": "deleted"}


_CONTENT_JOINS = ("LEFT JOIN media m ON m.id = pl.media_id "
                  "LEFT JOIN advertisers a ON a.id = m.advertiser_id "
                  "LEFT JOIN screens s ON s.id = pl.screen_id")


def _data_playtime(db: Session, days: int = 7, advertiser: Optional[str] = None,
                   date_from=None, date_to=None, media_id=None, screen=None, dow=None):
    where, p = _playlog_where(days, date_from, date_to, advertiser, media_id, screen, dow)
    total = db.execute(text(f"""
        SELECT COUNT(*) as plays,
               COALESCE(SUM(EXTRACT(EPOCH FROM (pl.ended_at - pl.started_at))), 0) as seconds
        FROM play_log pl {_CONTENT_JOINS}
        WHERE {where}
    """), p).fetchone()

    by_media = db.execute(text(f"""
        SELECT m.title, m.filename, a.name AS advertiser, COUNT(*) as plays,
               COALESCE(SUM(EXTRACT(EPOCH FROM (pl.ended_at - pl.started_at))), 0) as seconds
        FROM play_log pl {_CONTENT_JOINS}
        WHERE {where}
        GROUP BY m.id, m.title, m.filename, a.name
        ORDER BY plays DESC
    """), p).fetchall()

    by_advertiser = db.execute(text(f"""
        SELECT a.name AS advertiser, COUNT(*) as plays,
               COALESCE(SUM(EXTRACT(EPOCH FROM (pl.ended_at - pl.started_at))), 0) as seconds
        FROM play_log pl {_CONTENT_JOINS}
        WHERE {where} AND a.name IS NOT NULL
        GROUP BY a.id, a.name
        ORDER BY seconds DESC
    """), p).fetchall()

    return {
        "total": dict(total._mapping) if total else {"plays": 0, "seconds": 0},
        "by_media": [dict(r._mapping) for r in by_media],
        "by_advertiser": [dict(r._mapping) for r in by_advertiser],
    }


def _data_fillers(db: Session, days: int = 7, date_from=None, date_to=None,
                  media_id=None, screen=None, dow=None):
    """Отчёт по заглушкам: показы и эфирное время за период (всего и по каждому
    ролику) + справочно состав папки «Заглушки» в медиатеке."""
    where, p = _playlog_where(days, date_from, date_to, None, media_id, screen, dow)
    where += " AND COALESCE(m.is_filler, FALSE) = TRUE"

    air_total = db.execute(text(f"""
        SELECT COUNT(*) as plays,
               COALESCE(SUM(EXTRACT(EPOCH FROM (pl.ended_at - pl.started_at))), 0) as seconds
        FROM play_log pl {_CONTENT_JOINS}
        WHERE {where}
    """), p).fetchone()

    air_by_media = db.execute(text(f"""
        SELECT m.id AS media_id, m.title, m.filename, COUNT(*) as plays,
               COALESCE(SUM(EXTRACT(EPOCH FROM (pl.ended_at - pl.started_at))), 0) as seconds
        FROM play_log pl {_CONTENT_JOINS}
        WHERE {where}
        GROUP BY m.id, m.title, m.filename
        ORDER BY seconds DESC
    """), p).fetchall()

    library = db.execute(text("""
        SELECT id, title, filename, duration_seconds, status, review_status
        FROM media WHERE COALESCE(is_filler, FALSE) = TRUE
        ORDER BY title, filename
    """)).fetchall()
    lib_items = [dict(r._mapping) for r in library]
    lib_ready = [i for i in lib_items
                 if i["status"] == "ready" and i["review_status"] == "approved"]

    return {
        "air": {"plays": air_total.plays, "seconds": float(air_total.seconds or 0),
                "by_media": [dict(r._mapping) for r in air_by_media]},
        "library": {"count": len(lib_items), "ready_count": len(lib_ready),
                    "total_seconds": round(sum(i["duration_seconds"] or 0 for i in lib_ready), 1),
                    "items": lib_items},
    }


@router.get("/reports/fillers")
def report_fillers(days: int = Query(7),
                   date_from: Optional[str] = Query(None), date_to: Optional[str] = Query(None),
                   media_id: Optional[int] = Query(None), screen: Optional[str] = Query(None),
                   dow: Optional[int] = Query(None),
                   admin: dict = Depends(get_current_admin), db: Session = Depends(get_db)):
    return _data_fillers(db, days, date_from, date_to, media_id, screen, dow)


@router.get("/reports/playtime")
def report_playtime(days: int = Query(7), advertiser: Optional[str] = Query(None),
                     date_from: Optional[str] = Query(None), date_to: Optional[str] = Query(None),
                     media_id: Optional[int] = Query(None), screen: Optional[str] = Query(None),
                   dow: Optional[int] = Query(None),
                     admin: dict = Depends(get_current_admin), db: Session = Depends(get_db)):
    return _data_playtime(db, days, advertiser, date_from, date_to, media_id, screen, dow)


def _data_by_screen_advertiser(db: Session, days: int = 7, advertiser: Optional[str] = None,
                               date_from=None, date_to=None, media_id=None, screen=None, dow=None):
    where, p = _playlog_where(days, date_from, date_to, advertiser, media_id, screen, dow)
    rows = db.execute(text(f"""
        SELECT s.name AS screen, a.name AS advertiser, m.title, COUNT(*) as plays
        FROM play_log pl
        JOIN screens s ON s.id = pl.screen_id
        JOIN media m ON m.id = pl.media_id
        LEFT JOIN advertisers a ON a.id = m.advertiser_id
        WHERE {where}
        GROUP BY s.name, a.name, m.title
        ORDER BY plays DESC
    """), p).fetchall()
    return [dict(r._mapping) for r in rows]


@router.get("/reports/by-screen-advertiser")
def report_by_screen_advertiser(days: int = Query(7), advertiser: Optional[str] = Query(None),
                                 date_from: Optional[str] = Query(None), date_to: Optional[str] = Query(None),
                                 media_id: Optional[int] = Query(None), screen: Optional[str] = Query(None),
                   dow: Optional[int] = Query(None),
                                 admin: dict = Depends(get_current_admin), db: Session = Depends(get_db)):
    return _data_by_screen_advertiser(db, days, advertiser, date_from, date_to, media_id, screen, dow)


def _data_earnings(db: Session, days: int = 30, advertiser: Optional[str] = None,
                   date_from=None, date_to=None, media_id=None, screen=None, dow=None):
    where, p = _playlog_where(days, date_from, date_to, advertiser, media_id, screen, dow)
    rows = db.execute(text(f"""
        SELECT a.name AS advertiser, COUNT(*) as plays,
               COALESCE(SUM(EXTRACT(EPOCH FROM (pl.ended_at - pl.started_at))), 0) / 60.0 as minutes,
               a.price_per_minute
        FROM play_log pl
        JOIN media m ON m.id = pl.media_id
        JOIN advertisers a ON a.id = m.advertiser_id
        LEFT JOIN screens s ON s.id = pl.screen_id
        WHERE {where}
        GROUP BY a.id, a.name, a.price_per_minute
        ORDER BY minutes DESC
    """), p).fetchall()

    by_advertiser = []
    total_earnings = 0.0
    for r in rows:
        minutes = round(float(r.minutes or 0), 1)
        price = float(r.price_per_minute or 0)
        earnings = round(minutes * price, 2)
        total_earnings += earnings
        by_advertiser.append({
            "advertiser": r.advertiser, "plays": r.plays,
            "minutes": minutes, "price_per_minute": price, "earnings": earnings,
        })

    return {"by_advertiser": by_advertiser, "total_earnings": round(total_earnings, 2)}


@router.get("/reports/earnings")
def report_earnings(days: int = Query(30), advertiser: Optional[str] = Query(None),
                     date_from: Optional[str] = Query(None), date_to: Optional[str] = Query(None),
                     media_id: Optional[int] = Query(None), screen: Optional[str] = Query(None),
                   dow: Optional[int] = Query(None),
                     admin: dict = Depends(get_current_admin), db: Session = Depends(get_db)):
    return _data_earnings(db, days, advertiser, date_from, date_to, media_id, screen, dow)


def _playlog_where(days=7, date_from=None, date_to=None, advertiser=None,
                   media_id=None, screen=None, dow=None):
    """Собрать WHERE и параметры для журнала показов из белого списка фильтров.
    Если задан диапазон дат (date_from/date_to) — используется он; иначе — период
    в днях (days). Все значения передаются как bind-параметры (без вставки в SQL)."""
    where, params = [], {}
    if date_from or date_to:
        if date_from:
            where.append("pl.started_at >= CAST(:d_from AS timestamp) - INTERVAL '3 hours'")
            params["d_from"] = date_from
        if date_to:
            where.append("pl.started_at < CAST(:d_to AS timestamp) + INTERVAL '21 hours'")
            params["d_to"] = date_to
    else:
        where.append("pl.started_at > NOW() - (INTERVAL '1 day' * :days)")
        params["days"] = int(days or 7)
    if advertiser:
        where.append("a.name = :advertiser")
        params["advertiser"] = advertiser
    if media_id:
        where.append("pl.media_id = :media_id")
        params["media_id"] = int(media_id)
    if screen:
        where.append("s.name = :screen")
        params["screen"] = screen
    if dow:
        # День недели по МСК: 1 = понедельник … 7 = воскресенье (ISO)
        where.append("EXTRACT(ISODOW FROM pl.started_at + INTERVAL '3 hours') = :dow")
        params["dow"] = int(dow)
    return " AND ".join(where), params


def _data_playlog(db: Session, days: int = 7, advertiser: Optional[str] = None, limit: int = 500,
                  date_from: Optional[str] = None, date_to: Optional[str] = None,
                  media_id: Optional[int] = None, screen: Optional[str] = None, dow=None):
    where, params = _playlog_where(days, date_from, date_to, advertiser, media_id, screen, dow)
    params["limit"] = limit
    rows = db.execute(text(f"""
        SELECT to_char(pl.started_at + INTERVAL '3 hours', 'YYYY-MM-DD') as day,
               EXTRACT(ISODOW FROM pl.started_at + INTERVAL '3 hours')::int AS dow_num,
               to_char(pl.started_at + INTERVAL '3 hours', 'HH24:MI:SS') as time,
               s.name AS screen, COALESCE(m.title, pl.filename) AS media, a.name AS advertiser
        FROM play_log pl
        JOIN screens s ON s.id = pl.screen_id
        LEFT JOIN media m ON m.id = pl.media_id
        LEFT JOIN advertisers a ON a.id = m.advertiser_id
        WHERE {where}
        ORDER BY pl.started_at DESC
        LIMIT :limit
    """), params).fetchall()
    out = []
    for r in rows:
        d = dict(r._mapping)
        d["dow"] = DOW_RU[(d.pop("dow_num") or 1) - 1]
        out.append(d)
    return out


@router.get("/playlog")
def get_playlog(days: int = Query(7), advertiser: Optional[str] = Query(None),
                 limit: int = Query(500), date_from: Optional[str] = Query(None),
                 date_to: Optional[str] = Query(None), media_id: Optional[int] = Query(None),
                 screen: Optional[str] = Query(None),
                   dow: Optional[int] = Query(None),
                 admin: dict = Depends(get_current_admin), db: Session = Depends(get_db)):
    return _data_playlog(db, days, advertiser, limit, date_from, date_to, media_id, screen, dow)


def _data_playlog_by_advertiser(db: Session, days: int = 7, date_from=None, date_to=None,
                                media_id=None, screen=None, dow=None):
    where, params = _playlog_where(days, date_from, date_to, None, media_id, screen, dow)
    rows = db.execute(text(f"""
        SELECT a.name AS advertiser, COUNT(*) as plays,
               COUNT(DISTINCT DATE(pl.started_at)) as days,
               COUNT(DISTINCT pl.screen_id) as screens,
               to_char(MIN(pl.started_at) + INTERVAL '3 hours', 'YYYY-MM-DD HH24:MI') as first_play,
               to_char(MAX(pl.started_at) + INTERVAL '3 hours', 'YYYY-MM-DD HH24:MI') as last_play
        FROM play_log pl
        JOIN media m ON m.id = pl.media_id
        JOIN advertisers a ON a.id = m.advertiser_id
        JOIN screens s ON s.id = pl.screen_id
        WHERE {where}
        GROUP BY a.id, a.name
        ORDER BY plays DESC
    """), params).fetchall()
    return [dict(r._mapping) for r in rows]


@router.get("/playlog/by-advertiser")
def get_playlog_by_advertiser(days: int = Query(7), date_from: Optional[str] = Query(None),
                               date_to: Optional[str] = Query(None), media_id: Optional[int] = Query(None),
                               screen: Optional[str] = Query(None),
                   dow: Optional[int] = Query(None),
                               admin: dict = Depends(get_current_admin), db: Session = Depends(get_db)):
    return _data_playlog_by_advertiser(db, days, date_from, date_to, media_id, screen, dow)


# ─── Экспорт отчётов в Excel/PDF ─────────────────────────────────────────────
# DejaVu (уже стоит в Dockerfile ради этого) — нужен для кириллицы в PDF,
# у reportlab встроенные шрифты кириллицу не поддерживают.
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

try:
    pdfmetrics.registerFont(TTFont('DejaVuSans', '/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf'))
    pdfmetrics.registerFont(TTFont('DejaVuSans-Bold', '/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf'))
except Exception:
    log.warning("Не удалось зарегистрировать DejaVuSans для PDF — кириллица может не отобразиться")


def _cell(v):
    if v is None:
        return "—"
    if isinstance(v, datetime):
        # В БД метки времени в UTC; во всех выгрузках показываем МСК
        return (v + timedelta(hours=3)).strftime("%Y-%m-%d %H:%M")
    if hasattr(v, "strftime"):  # чистая дата — без сдвига
        return v.strftime("%Y-%m-%d")
    return v


def _build_xlsx(sections):
    """sections: список (title, columns, rows). Оформление: жирная шапка с
    заливкой, закреплённая первая строка, автофильтр, ширины столбцов по
    содержимому (наименования не обрезаются)."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    wb = Workbook()
    wb.remove(wb.active)
    head_font = Font(bold=True, color="FFFFFF")
    head_fill = PatternFill("solid", fgColor="2B2F36")
    thin = Side(style="thin", color="B7BCC4")
    box = Border(left=thin, right=thin, top=thin, bottom=thin)
    used_names = set()
    invalid_chars = r'[]:*?/\\'
    for title, columns, rows in sections:
        safe_title = "".join(c for c in (title or "Отчёт") if c not in invalid_chars)
        name = safe_title[:31] or "Отчёт"
        base, i = name, 1
        while name in used_names:
            i += 1
            name = f"{base[:28]}~{i}"
        used_names.add(name)
        ws = wb.create_sheet(name)
        ws.append(list(columns))
        for c in ws[1]:
            c.font = head_font
            c.fill = head_fill
            c.alignment = Alignment(vertical="center")
            c.border = box
        for row in rows:
            ws.append([_cell(c) for c in row])
        if not rows:
            # Пустой раздел — явная пометка вместо голой шапки
            ws.append(["Нет данных за выбранный период"] + [""] * (len(columns) - 1))
            ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max(1, len(columns)))
            ws.cell(row=2, column=1).font = Font(italic=True, color="808690")
        # Рамки вокруг всех ячеек данных — отчёт выглядит законченным блоком
        for r in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=len(columns)):
            for c in r:
                c.border = box
        # Ширины столбцов: по самому длинному значению (шапка или данные),
        # с разумным потолком — длинное содержимое переносится внутри ячейки
        for ci in range(1, len(columns) + 1):
            longest = max([len(str(columns[ci - 1]))] +
                          [len(str(_cell(r[ci - 1]))) for r in rows[:200]] or [0])
            ws.column_dimensions[get_column_letter(ci)].width = min(52, max(10, longest + 2))
        if rows:
            last_col = get_column_letter(len(columns))
            ws.auto_filter.ref = f"A1:{last_col}{len(rows) + 1}"
        ws.freeze_panes = "A2"
        # Серую «бесконечную» сетку вокруг отчёта не показываем —
        # виден только сам отчёт в рамках
        ws.sheet_view.showGridLines = False
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ─── Занятость инвентаря: сколько эфирного времени продано ─────────────────

@router.get("/reports/inventory")
def report_inventory(admin: dict = Depends(get_current_admin), db: Session = Depends(get_db)):
    """
    Загрузка недельного шаблона: сколько из 168 часов недели у каждого экрана
    закрыто слотом (с учётом трёхуровневой резолюции экран > группа > сеть;
    слот «весь день» закрывает все 24 часа). Незакрытые часы — заглушки,
    то есть непроданный инвентарь.
    """
    screens = db.execute(text(
        "SELECT id, name, group_id FROM screens ORDER BY name")).fetchall()
    slots = db.execute(text(
        "SELECT screen_id, group_id, day_of_week, hour FROM schedule_slots")).fetchall()

    # индексы слотов по уровням
    scr_slots, grp_slots, net_slots = {}, {}, set()
    for s in slots:
        key = (s.day_of_week, s.hour)  # hour None = весь день
        if s.screen_id is not None:
            scr_slots.setdefault(s.screen_id, set()).add(key)
        elif s.group_id is not None:
            grp_slots.setdefault(s.group_id, set()).add(key)
        else:
            net_slots.add(key)

    def busy_hours(sid, gid):
        merged = net_slots | grp_slots.get(gid, set()) | scr_slots.get(sid, set())
        count = 0
        for dow in range(7):
            all_day = (dow, None) in merged
            for hr in range(24):
                if all_day or (dow, hr) in merged:
                    count += 1
        return count

    by_screen = []
    total_busy = 0
    for s in screens:
        busy = busy_hours(s.id, s.group_id)
        total_busy += busy
        by_screen.append({"screen_id": s.id, "name": s.name,
                          "hours_busy": busy, "hours_total": 168,
                          "fill_pct": round(100.0 * busy / 168, 1)})

    total_hours = 168 * len(screens)
    return {
        "screens_total": len(screens),
        "network_fill_pct": round(100.0 * total_busy / total_hours, 1) if total_hours else 0,
        "hours_busy_total": total_busy,
        "hours_total": total_hours,
        "by_screen": by_screen,
    }


# ─── Отчёт для ФАС (38-ФЗ): декларация → решение модератора → факт показов ──

def _data_fas(db: Session, days: int = 90):
    from routers.media import AD_CATEGORIES
    rows = db.execute(text("""
        SELECT m.id, m.title, m.category, m.age_rating,
               m.disclaimer_text, m.license_number,
               m.review_status, m.reviewed_by, m.reviewed_at, m.reject_reason,
               a.name AS advertiser,
               COUNT(pl.id) AS plays,
               COUNT(DISTINCT pl.screen_id) AS screens,
               MIN(pl.started_at) AS first_shown,
               MAX(pl.started_at) AS last_shown
        FROM media m
        LEFT JOIN advertisers a ON a.id = m.advertiser_id
        LEFT JOIN play_log pl ON pl.media_id = m.id
             AND pl.started_at > NOW() - (INTERVAL '1 day' * :days)
        WHERE m.status = 'ready'
        GROUP BY m.id, a.name
        ORDER BY a.name NULLS LAST, m.title
    """), {"days": days}).fetchall()

    items = []
    for r in rows:
        cat = AD_CATEGORIES.get(r.category or "other", {})
        items.append({
            "id": r.id, "title": r.title, "advertiser": r.advertiser,
            "category": r.category, "category_label": cat.get("label", r.category),
            "age_rating": r.age_rating,
            "disclaimer_text": r.disclaimer_text, "license_number": r.license_number,
            "review_status": r.review_status, "reviewed_by": r.reviewed_by,
            "reviewed_at": r.reviewed_at, "reject_reason": r.reject_reason,
            "plays": r.plays, "screens": r.screens,
            "first_shown": r.first_shown, "last_shown": r.last_shown,
        })

    decisions = db.execute(text("""
        SELECT created_at, title, detail, actor FROM audit_log
        WHERE event_type = 'moderation'
          AND created_at > NOW() - (INTERVAL '1 day' * :days)
        ORDER BY created_at DESC
    """), {"days": days}).fetchall()

    return {"days": days, "items": items,
            "decisions": [dict(r._mapping) for r in decisions]}


@router.get("/reports/fas")
def report_fas(days: int = Query(90), admin: dict = Depends(get_current_admin),
               db: Session = Depends(get_db)):
    return _data_fas(db, days)


@router.get("/reports/fas.pdf")
def report_fas_pdf(days: int = Query(90), admin: dict = Depends(get_current_admin),
                   db: Session = Depends(get_db)):
    """PDF-отчёт для регулятора: что показывалось, под какой декларацией,
    кто и когда одобрил. Вместе с журналом показов — доказательная база."""
    d = _data_fas(db, days)

    def _dt(v):
        # UTC из БД → МСК
        return (v + timedelta(hours=3)).strftime("%d.%m.%Y %H:%M") if v else "—"

    st_ru = {"approved": "одобрен", "pending": "на модерации", "rejected": "отклонён"}
    main_rows = []
    for it in d["items"]:
        decision = st_ru.get(it["review_status"], it["review_status"])
        if it["reviewed_by"]:
            decision += f" ({it['reviewed_by']}, {_dt(it['reviewed_at'])})"
        period = f"{_dt(it['first_shown'])} — {_dt(it['last_shown'])}" if it["plays"] else "—"
        # Текст не обрезаем: ячейки PDF переносят строки (Paragraph в _build_pdf)
        main_rows.append([
            it["title"] or "", it["advertiser"] or "—",
            it["category_label"] or "", it["age_rating"] or "—",
            decision, it["plays"], period,
        ])

    reg_rows = [[it["title"] or "",
                 it["disclaimer_text"] or "—",
                 it["license_number"] or "—"]
                for it in d["items"] if it["disclaimer_text"] or it["license_number"]]

    dec_rows = [[_dt(r["created_at"]), r["title"] or "",
                 r["detail"] or "", r["actor"]]
                for r in d["decisions"]]

    sections = [
        (f"Реклама в эфире за {days} дн.: декларация и показы",
         ["Ролик", "Рекламодатель", "Категория (38-ФЗ)", "Маркировка",
          "Решение модератора", "Показов", "Период показа (МСК)"], main_rows),
        ("Обязательные предупреждения и лицензии",
         ["Ролик", "Текст предупреждения", "№ лицензии"], reg_rows),
        ("Журнал решений модерации",
         ["Дата (МСК)", "Решение", "Детали", "Модератор"], dec_rows),
    ]
    buf = _build_pdf(sections, doc_title=f"Отчёт о соответствии закону «О рекламе» (38-ФЗ) — {days} дн.")
    return StreamingResponse(buf, media_type="application/pdf", headers={
        "Content-Disposition": 'attachment; filename="fas_report.pdf"'})


def _build_pdf(sections, doc_title="Отчёт"):
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, title=doc_title)
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('TitleRu', parent=styles['Title'], fontName='DejaVuSans-Bold')
    heading_style = ParagraphStyle('HeadingRu', parent=styles['Heading2'], fontName='DejaVuSans-Bold')
    normal_style = ParagraphStyle('NormalRu', parent=styles['Normal'], fontName='DejaVuSans')

    # Текст в ячейках оборачиваем в Paragraph: длинные значения ПЕРЕНОСЯТСЯ
    # на новую строку, а не обрезаются; таблица растягивается на ширину листа.
    cell_style = ParagraphStyle('CellRu', parent=styles['Normal'],
                                fontName='DejaVuSans', fontSize=8, leading=10)
    head_style = ParagraphStyle('CellHeadRu', parent=cell_style,
                                fontName='DejaVuSans-Bold', textColor=colors.white)

    elements = [Paragraph(doc_title, title_style), Spacer(1, 12)]
    avail_width = A4[0] - doc.leftMargin - doc.rightMargin
    for title, columns, rows in sections:
        elements.append(Paragraph(title, heading_style))
        elements.append(Spacer(1, 4))
        if rows:
            data = [[Paragraph(str(c), head_style) for c in columns]]
            for row in rows:
                data.append([Paragraph(str(_cell(c)), cell_style) for c in row])
            # Ширины столбцов — по длине содержимого (шапка + первые строки),
            # с минимумом, чтобы узкие столбцы не схлопывались
            # Каждому столбцу — гарантированный минимум в пунктах под самое
            # длинное слово заголовка (иначе шапка ломается по буквам:
            # «Рекла-модат-ель»); остаток ширины делится по длине содержимого.
            mins, weights = [], []
            for ci in range(len(columns)):
                head_word = max(len(w) for w in str(columns[ci]).split())
                mins.append(head_word * 5.8 + 12)          # ~8pt bold + отступы
                sample = [len(str(_cell(r[ci]))) for r in rows[:30]]
                weights.append(max(4, min(60, max(sample))))
            extra = avail_width - sum(mins)
            if extra > 0:
                wsum = sum(weights)
                col_widths = [mins[i] + extra * weights[i] / wsum
                              for i in range(len(columns))]
            else:  # столбцов слишком много — ужимаем пропорционально минимумам
                col_widths = [avail_width * m / sum(mins) for m in mins]
            t = Table(data, repeatRows=1, colWidths=col_widths)
            t.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2b2f36')),
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('LEFTPADDING', (0, 0), (-1, -1), 4),
                ('RIGHTPADDING', (0, 0), (-1, -1), 4),
                ('TOPPADDING', (0, 0), (-1, -1), 3),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f2f2f2')]),
            ]))
            elements.append(t)
        else:
            elements.append(Paragraph("Нет данных", normal_style))
        elements.append(Spacer(1, 16))
    doc.build(elements)
    buf.seek(0)
    return buf


def _section_table(key: str, db: Session, days: int):
    if key == "offline":
        rows = _data_offline_screens(db)
        return ("Неработающие экраны", ["Экран", "Город", "Последний раз на связи"],
                [[r["name"], r["city"], r["last_seen"]] for r in rows])
    if key == "downtime":
        rows = _data_downtime(db)
        return ("Простой экранов", ["Экран", "Простой, %"],
                [[r["name"], r["downtime_pct"]] for r in rows])
    if key == "lowdisk":
        rows = _data_low_disk(db)
        return ("Память < 15%", ["Экран", "Свободно, ГБ", "Всего, ГБ", "Свободно, %"],
                [[r["name"], r["disk_free_gb"], r["disk_total_gb"], r["free_pct"]] for r in rows])
    if key == "broken":
        rows = _data_broken_media(db)
        return ("Нерабочие ролики", ["Ролик", "Рекламодатель", "Ошибок", "Последняя ошибка"],
                [[r["title"] or r["filename"], r["advertiser"], r["error_count"], r["last_error"]] for r in rows])
    if key == "by_media":
        rows = _data_playtime(db, days)["by_media"]
        return ("Время показа по роликам", ["Ролик", "Рекламодатель", "Повторов", "Время показа, сек"],
                [[r["title"] or r["filename"], r["advertiser"], r["plays"], round(r["seconds"])] for r in rows])
    if key == "by_advertiser":
        rows = _data_playtime(db, days)["by_advertiser"]
        return ("Время показа по рекламодателям", ["Рекламодатель", "Повторов", "Время показа, сек"],
                [[r["advertiser"], r["plays"], round(r["seconds"])] for r in rows])
    if key == "by_screen":
        rows = _data_by_screen_advertiser(db, days)
        return ("Показы по экранам", ["Экран", "Рекламодатель", "Ролик", "Повторов"],
                [[r["screen"], r["advertiser"], r["title"], r["plays"]] for r in rows])
    if key == "sessions":
        rows = _data_sessions(db, days)
        return ("Входы пользователей", ["Пользователь", "Роль", "Вход", "Секунд"],
                [[r["username"], r["role"], r["login_at"], round(r["seconds"] or 0)] for r in rows])
    if key == "validity":
        rows = _data_validity(db)
        return ("Сроки действия роликов", ["Ролик", "Рекламодатель", "Показывать с", "Показывать по", "Статус"],
                [[r["title"], r["advertiser"], r["valid_from"], r["valid_until"], r["validity_status"]] for r in rows])
    if key == "conn_loss":
        rows = _data_connection_losses(db, days)
        return ("Потери связи с ПК", ["Экран", "Связь пропала", "Восстановлена", "Секунд", "Статус"],
                [[r["screen"], r["lost_at"], r["restored_at"], round(r["seconds"] or 0),
                  "нет связи" if r["ongoing"] else "восстановлена"] for r in rows])
    if key == "earnings":
        rows = _data_earnings(db, days)["by_advertiser"]
        return ("Заработок по рекламодателям", ["Рекламодатель", "Показов", "Минут эфира", "Цена ₽/мин", "Заработано ₽"],
                [[r["advertiser"], r["plays"], r["minutes"], r["price_per_minute"], r["earnings"]] for r in rows])
    return (key, ["—"], [])


@router.get("/reports/export.xlsx")
def export_reports_xlsx(days: int = Query(7), sections: str = Query(""),
                         admin: dict = Depends(get_current_admin), db: Session = Depends(get_db)):
    keys = [k for k in sections.split(",") if k]
    if not keys:
        raise HTTPException(status_code=400, detail="Не выбраны разделы для экспорта")
    secs = [_section_table(k, db, days) for k in keys]
    buf = _build_xlsx(secs)
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                              headers={"Content-Disposition": 'attachment; filename="ds_report.xlsx"'})


@router.get("/reports/export.pdf")
def export_reports_pdf(days: int = Query(7), sections: str = Query(""),
                        admin: dict = Depends(get_current_admin), db: Session = Depends(get_db)):
    keys = [k for k in sections.split(",") if k]
    if not keys:
        raise HTTPException(status_code=400, detail="Не выбраны разделы для экспорта")
    secs = [_section_table(k, db, days) for k in keys]
    buf = _build_pdf(secs, doc_title="Отчёты Digital Signage")
    return StreamingResponse(buf, media_type="application/pdf",
                              headers={"Content-Disposition": 'attachment; filename="ds_report.pdf"'})


@router.get("/playlog/export.xlsx")
def export_playlog_xlsx(days: int = Query(7), advertiser: Optional[str] = Query(None),
                         date_from: Optional[str] = Query(None), date_to: Optional[str] = Query(None),
                         media_id: Optional[int] = Query(None), screen: Optional[str] = Query(None),
                   dow: Optional[int] = Query(None),
                         admin: dict = Depends(get_current_admin), db: Session = Depends(get_db)):
    rows = _data_playlog(db, days, advertiser, 5000, date_from, date_to, media_id, screen, dow)
    columns = ["Дата", "День", "Время", "Экран", "Ролик", "Рекламодатель"]
    data_rows = [[r["day"], r["dow"], r["time"], r["screen"], r["media"], r["advertiser"]] for r in rows]
    buf = _build_xlsx([("Журнал показов", columns, data_rows)])
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                              headers={"Content-Disposition": 'attachment; filename="playlog.xlsx"'})


@router.get("/playlog/export.pdf")
def export_playlog_pdf(days: int = Query(7), advertiser: Optional[str] = Query(None),
                        date_from: Optional[str] = Query(None), date_to: Optional[str] = Query(None),
                        media_id: Optional[int] = Query(None), screen: Optional[str] = Query(None),
                   dow: Optional[int] = Query(None),
                        admin: dict = Depends(get_current_admin), db: Session = Depends(get_db)):
    rows = _data_playlog(db, days, advertiser, 2000, date_from, date_to, media_id, screen, dow)
    columns = ["Дата", "День", "Время", "Экран", "Ролик", "Рекламодатель"]
    data_rows = [[r["day"], r["dow"], r["time"], r["screen"], r["media"], r["advertiser"]] for r in rows]
    buf = _build_pdf([("Журнал показов", columns, data_rows)], doc_title="Журнал показов")
    return StreamingResponse(buf, media_type="application/pdf",
                             headers={"Content-Disposition": 'attachment; filename="playlog.pdf"'})


@router.post("/reports/builder/export.xlsx")
def builder_export_xlsx(body: dict = Body(...), admin: dict = Depends(get_current_admin),
                         db: Session = Depends(get_db)):
    title, columns, rows = _builder_run_data(body, db)
    buf = _build_xlsx([(title, columns, rows)])
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                              headers={"Content-Disposition": 'attachment; filename="ds_custom_report.xlsx"'})


@router.post("/reports/builder/export.pdf")
def builder_export_pdf(body: dict = Body(...), admin: dict = Depends(get_current_admin),
                        db: Session = Depends(get_db)):
    title, columns, rows = _builder_run_data(body, db)
    buf = _build_pdf([(title, columns, rows)], doc_title=title)
    return StreamingResponse(buf, media_type="application/pdf",
                              headers={"Content-Disposition": 'attachment; filename="ds_custom_report.pdf"'})
